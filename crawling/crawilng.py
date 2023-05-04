from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from datetime import datetime

html = urlopen("https://www.lawtimes.co.kr/Legal-News/Legal-News-List?page=1")
bsobj = BeautifulSoup(html , "html.parser")

utl_intro = 'https://www.lawtimes.co.kr/'

article_div = bsobj.find("div", "tabs4")
articles = article_div.find_all("article", "article-list")

wb = openpyxl.Workbook()
sheet = wb.active

basic = ['기사 제목' , 'URL', '날짜']
sheet.append(basic)

for i in articles:
  
  article_list = []
  
  article = i.find("a")
  article_info = article.find("div",'article-txt')
  
  title = article_info.find('h4')
  article_list.append(title.text)
  
  url = utl_intro+article['href']
  article_list.append(url)
  
  time = article_info.find('time')
  article_list.append(time.text)
  
  sheet.append(article_list)
  
today = datetime.today().strftime("%Y%m%d_%H시%M분") 

wb.save(f'./기사/{today}.xlsx')
wb.close

